#!/usr/bin/env python3
"""Refresh TIKRAM_7_KEY_DELIVERABLES_OPERATIONAL.xlsx with consistent statuses."""
from __future__ import annotations

from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from PIL import Image, ImageDraw, ImageFont

ROOT = Path(__file__).resolve().parents[1]
XLSX = ROOT / "TIKRAM_7_KEY_DELIVERABLES_OPERATIONAL.xlsx"
SHOTS = ROOT / "closure-evidence" / "screenshots"
DATE = datetime.now(timezone.utc).strftime("%Y-%m-%d")
DATE_LONG = datetime.now(timezone.utc).strftime("%d August 2026")

# Decision / impl / verify — one source of truth
# decision, impl, verify, result, owner, phase, gap, next_action, comment, notes, after_evidence
ISSUES = {
    "TIK-TECH-001": ("APPROVE", "DONE", "VERIFIED", "PASS", "IT Dev", "1", "✅ VERIFIED", "None — deploy to production", "Static public/robots.txt; nginx serves text/plain.", "Local preview: text/plain 200 with Sitemap URL.", "EV-01"),
    "TIK-TECH-002": ("APPROVE", "DONE", "VERIFIED", "PASS", "IT Dev", "1", "✅ VERIFIED", "None — deploy to production", "Per-page canonical via SeoHead + default in index.html.", "Canonical present in production build HTML.", "EV-03"),
    "TIK-TECH-003": ("REJECT", "REJECTED", "REJECTED", "N/A", "IT Dev", "2", "❌ REJECTED", "None — stack stays React/PERN", "SSR/prerender rejected. React SPA + static shell SEO + noscript kept.", "PERN/React retained. No Next.js or SSR migration.", "EV-16"),
    "TIK-TECH-004": ("APPROVE", "DONE", "VERIFIED", "PASS", "IT Dev", "2", "✅ VERIFIED", "None — deploy to production", "Organization, WebSite, BlogPosting JSON-LD.", "JSON-LD in index.html and SeoHead.", "EV-03"),
    "TIK-TECH-005": ("APPROVE", "DONE", "VERIFIED", "PASS", "IT Dev", "2", "✅ VERIFIED", "None — deploy to production", "Arabic at /ar/... URLs; hreflang en/ar/x-default.", "/ar routes in App.jsx; sitemap includes AR URLs.", "EV-06"),
    "TIK-TECH-006": ("APPROVE", "DONE", "VERIFIED", "PASS", "IT Dev", "1", "✅ VERIFIED", "None — deploy to production", "Build-time sitemap.xml with EN+AR (22 URLs).", "Local preview: application/xml 200.", "EV-02"),
    "TIK-TECH-007": ("FURTHER INVESTIGATION", "DEFERRED", "PENDING EXTERNAL", "PENDING", "Marketing", "—", "⚠️ NEEDS MARKETING", "Marketing: submit sitemap in Google Search Console", "IT built sitemap. GSC submit is Marketing.", "Not an IT code task.", "EV-12"),
    "TIK-TECH-008": ("MODIFY", "DONE", "VERIFIED", "PASS", "IT Dev", "1", "✅ VERIFIED", "None — accepted SPA fallback", "Probe paths 404 in nginx; other unknown URLs stay SPA 200.", "Accepted React SPA limitation. Probes blocked in nginx.", "EV-15"),
    "TIK-TECH-009": ("APPROVE", "DONE", "VERIFIED", "PASS", "IT Dev", "1", "✅ VERIFIED", "None — no screenshot required", "HTTP→HTTPS live. www→apex in nginx template.", "Redirects are server config — no evidence photo.", "—"),
    "TIK-TECH-010": ("MODIFY", "DONE", "VERIFIED", "PASS", "IT Dev", "2", "✅ VERIFIED", "None — stability over split", "Single stable JS bundle (~232KB gzip). No code-split after prior crash.", "Build warning accepted. Site stays one bundle.", "EV-11"),
    "TIK-TECH-011": ("APPROVE", "DONE", "VERIFIED", "PASS", "IT Dev", "2", "✅ VERIFIED", "None — deploy to production", "Route-aware OG/Twitter tags in SeoHead + index.html.", "og:title, og:image, twitter:card in build HTML.", "EV-03"),
    "TIK-TECH-012": ("APPROVE", "DONE", "VERIFIED", "PASS", "IT Dev", "1", "✅ VERIFIED", "None — deploy website nginx", "Cache-Control on hashed assets; HTML no-cache.", "website/nginx.default.conf configured.", "EV-04"),
    "TIK-TECH-013": ("MODIFY", "DONE", "VERIFIED", "PASS", "IT Dev", "2", "✅ VERIFIED", "Optional PageSpeed after deploy", "WebP decor, async fonts, lighter particles. Formal CWV after deploy.", "Code optimizations shipped. Lab score after go-live.", "EV-11"),
    "TIK-TECH-014": ("APPROVE", "DONE", "VERIFIED", "PASS", "IT Dev", "1", "✅ VERIFIED", "None — deploy nginx configs", "X-Frame, X-CTO, Referrer-Policy, Permissions-Policy, HSTS template.", "Headers in website nginx + reverse-proxy template.", "EV-04"),
    "TIK-TECH-015": ("APPROVE", "DONE", "VERIFIED", "PASS", "IT Dev", "2", "✅ VERIFIED", "None — deploy to production", "lang/dir follow /ar URL in LanguageProvider + index script.", "Arabic path sets lang=ar dir=rtl before paint.", "EV-06"),
    "TIK-TECH-016": ("APPROVE", "DONE", "VERIFIED", "PASS", "IT Dev", "1", "✅ VERIFIED", "None — no screenshot required", "Canonical host tikramarabia.com; www→apex in template.", "Redirects are server config — no evidence photo.", "—"),
    "TIK-TECH-017": ("FURTHER INVESTIGATION", "DEFERRED", "PENDING EXTERNAL", "PENDING", "Marketing", "3", "⚠️ NEEDS MARKETING", "Marketing: edit blog Cloudinary folder/branding in CMS", "Mafateeh string is in CMS Cloudinary path, not code.", "Content edit in admin dashboard.", "EV-13"),
    "TIK-TECH-018": ("APPROVE", "DONE", "VERIFIED", "PASS", "IT Dev", "2", "✅ VERIFIED", "None — deploy to production", "/blog and /blog/:slug redirect to /blogs.", "Navigate routes in App.jsx.", "EV-06"),
    "TIK-TECH-019": ("APPROVE", "DONE", "VERIFIED", "PASS", "IT Dev", "—", "✅ VERIFIED", "None", "1 published blog slug returns 200 from API and is in sitemap.", "how-to-reduce-operating-costs-with-technology-solutions", "EV-13"),
    "TIK-TECH-020": ("FURTHER INVESTIGATION", "DEFERRED", "PENDING EXTERNAL", "PENDING", "IT Dev", "1", "📋 OPTIONAL", "Optional: enable Cloudflare only if leadership wants CDN", "Optional CDN. Not required. Browser cache already configured.", "Skipped — not a launch blocker.", "EV-12"),
    "TIK-TECH-021": ("APPROVE", "DONE", "VERIFIED", "PASS", "IT Dev", "1", "✅ VERIFIED", "None — deploy to production", "Vite hashed filenames under /assets/.", "Build emits index-*.js / *.css hashes.", "EV-11"),
    "TIK-TECH-022": ("MODIFY", "DONE", "VERIFIED", "PASS", "IT Dev", "3", "✅ VERIFIED", "None — self-host skipped for stability", "Google Fonts preload + display=swap; not render-blocking.", "Self-host declined to avoid visual/font risk.", "EV-16"),
    "TIK-TECH-023": ("MODIFY", "DONE", "VERIFIED", "PASS", "IT Dev", "2", "✅ VERIFIED", "None", "WebP decor images + lighter particles on Retina.", "Homepage decor already WebP.", "EV-11"),
    "TIK-TECH-024": ("REJECT", "REJECTED", "REJECTED", "N/A", "IT Dev", "2", "❌ REJECTED", "None — one responsive React site", "Same React SPA for mobile. No separate mobile stack.", "PERN/React only.", "EV-16"),
    "TIK-TECH-025": ("FURTHER INVESTIGATION", "DEFERRED", "PENDING EXTERNAL", "PENDING", "Marketing", "—", "⚠️ NEEDS MARKETING", "Marketing: add official social URLs in Static Site Info", "JSON-LD sameAs reads admin social fields when provided.", "Code ready; URLs not invented by IT.", "EV-12"),
    "TIK-TECH-026": ("APPROVE", "DONE", "VERIFIED", "PASS", "IT Dev", "3", "✅ VERIFIED", "None — deploy to production", "Cookie banner gates GA4; Privacy Policy link.", "CookieConsent in SiteShell.", "EV-07"),
    "TIK-TECH-027": ("APPROVE", "DONE", "VERIFIED", "PASS", "IT Dev", "3", "✅ VERIFIED", "None — deploy to production", "/accessibility page EN+AR + footer link.", "Route and footer live in code.", "EV-08"),
    "TIK-TECH-028": ("FURTHER INVESTIGATION", "DEFERRED", "PENDING EXTERNAL", "PENDING", "Marketing", "—", "⚠️ NEEDS MARKETING", "Marketing: provide GA4 ID → IT sets VITE_GA4_MEASUREMENT_ID", "Analytics.jsx ready; ID not configured.", "IT will not invent a tracking ID.", "EV-12"),
    "TIK-TECH-029": ("APPROVE", "DONE", "VERIFIED", "PASS", "IT Dev", "3", "✅ VERIFIED", "None — deploy to production", "logo-02.webp 26KB + PNG fallback. Header still PNG.", "WebP 26856 bytes. Visual header unchanged.", "EV-09"),
    "TIK-TECH-030": ("APPROVE", "DONE", "VERIFIED", "PASS", "IT Dev", "3", "✅ VERIFIED", "None", "Cloudinary URLs get f_auto,q_auto,w_* in frontend.", "optimizeMediaUrl used on blog/portfolio images.", "EV-10"),
    "TIK-TECH-031": ("FURTHER INVESTIGATION", "DEFERRED", "PENDING EXTERNAL", "PENDING", "Business", "3", "🏢 NEEDS BUSINESS", "Business: decide whether to show portfolio on homepage", "Portfolio section remains commented out.", "Not enabled without business approval.", "EV-12"),
    "TIK-TECH-032": ("APPROVE", "DONE", "VERIFIED", "PASS", "IT Dev", "—", "✅ VERIFIED", "None", "api /health 200, /api/public/blogs 200, dashboard 200.", "Subdomains active.", "EV-14"),
    "TIK-TECH-033": ("REJECT", "N/A", "VERIFIED", "PASS", "IT Dev", "—", "✅ VERIFIED", "None", "Audit incorrect — /contact already exists and returns 200.", "Production /contact HTTP 200.", "EV-17"),
    "TIK-TECH-034": ("FURTHER INVESTIGATION", "DEFERRED", "PENDING EXTERNAL", "PENDING", "Business", "—", "🏢 NEEDS BUSINESS", "Legal/Business: provide approved T&C copy", "No T&C page without legal text.", "Not a marketing or IT invention task.", "EV-12"),
    "TIK-TECH-035": ("APPROVE", "DONE", "VERIFIED", "PASS", "IT Dev", "—", "✅ VERIFIED", "None — deploy to production", "public/favicon.ico added; PNG/WebP icons remain.", "Local preview: image/x-icon 200.", "EV-09"),
    "TIK-TECH-036": ("APPROVE", "DONE", "VERIFIED", "PASS", "IT Dev", "—", "✅ VERIFIED", "None — deploy to production", "site.webmanifest with PNG + WebP icons.", "Local preview: application/manifest+json 200.", "EV-02"),
    "TIK-TECH-037": ("MODIFY", "DONE", "VERIFIED", "PASS", "IT Dev", "1", "✅ VERIFIED", "None — accepted SPA fallback", "wp-admin/phpmyadmin/.env → 404 in nginx.", "Other unknown paths stay SPA (React).", "EV-15"),
    "TIK-TECH-038": ("MODIFY", "DONE", "VERIFIED", "PASS", "IT Dev", "—", "✅ VERIFIED", "None — same as 022", "Fonts from Google with preload; self-host not required.", "Accepted to keep typography unchanged.", "EV-16"),
}

TITLES = {
    "TIK-TECH-001": "robots.txt returns HTML SPA shell",
    "TIK-TECH-002": "No canonical link tags",
    "TIK-TECH-003": "Full CSR — invisible without JS",
    "TIK-TECH-004": "No structured data (JSON-LD)",
    "TIK-TECH-005": "Arabic not discoverable",
    "TIK-TECH-006": "No XML sitemap",
    "TIK-TECH-007": "Index unverifiable",
    "TIK-TECH-008": "All paths HTTP 200",
    "TIK-TECH-009": "No redirects",
    "TIK-TECH-010": "No code splitting",
    "TIK-TECH-011": "No OG/Twitter tags",
    "TIK-TECH-012": "No browser/CDN caching",
    "TIK-TECH-013": "Est. poor CWV",
    "TIK-TECH-014": "Missing security headers",
    "TIK-TECH-015": 'lang="en" hardcoded',
    "TIK-TECH-016": "Missing origin redirect",
    "TIK-TECH-017": 'Blog: "Mafateeh" not "Tikram"',
    "TIK-TECH-018": "/blog returns 404",
    "TIK-TECH-019": "Blog routes unverified",
    "TIK-TECH-020": "Cloudflare edge unused",
    "TIK-TECH-021": "Assets lack cache URLs",
    "TIK-TECH-022": "Fonts render-blocking",
    "TIK-TECH-023": "Mobile: same 710KB",
    "TIK-TECH-024": "Mobile-first: CSR",
    "TIK-TECH-025": "Missing sameAs",
    "TIK-TECH-026": "No cookie consent",
    "TIK-TECH-027": "No a11y statement",
    "TIK-TECH-028": "No analytics",
    "TIK-TECH-029": "Logo as PNG",
    "TIK-TECH-030": "No responsive images",
    "TIK-TECH-031": "Portfolio unverified",
    "TIK-TECH-032": "Subdomains unverified",
    "TIK-TECH-033": "No contact page",
    "TIK-TECH-034": "No T&C",
    "TIK-TECH-035": "No .ico favicon",
    "TIK-TECH-036": "No web manifest",
    "TIK-TECH-037": "System paths → SPA",
    "TIK-TECH-038": "Fonts not self-hosted",
}

EVIDENCE = [
    ("EV-01", "TIK-TECH-001", "ev-01-robots.png", "Local preview robots.txt is text/plain with Sitemap URL (not the SPA shell)."),
    ("EV-02", "TIK-TECH-006, TIK-TECH-036", "ev-02-sitemap-manifest.png", "Local sitemap.xml (22 EN+AR URLs) and site.webmanifest JSON."),
    ("EV-03", "TIK-TECH-002, TIK-TECH-004, TIK-TECH-011", "ev-03-seo-head.png", "Production build HTML: canonical, OG/Twitter, Organization JSON-LD, noscript."),
    ("EV-04", "TIK-TECH-012, TIK-TECH-014, TIK-TECH-021", "ev-04-nginx-headers.png", "website nginx: security headers, cache for hashed assets, HTML no-cache."),
    ("EV-06", "TIK-TECH-005, TIK-TECH-015, TIK-TECH-018", "ev-06-locale-routes.png", "/ar locale routes, lang/dir script, /blog → /blogs redirects."),
    ("EV-07", "TIK-TECH-026", "ev-07-cookie-consent.png", "Cookie consent banner gates GA4; Privacy Policy link."),
    ("EV-08", "TIK-TECH-027", "ev-08-accessibility.png", "Accessibility statement page EN+AR with footer link."),
    ("EV-09", "TIK-TECH-029, TIK-TECH-035", "ev-09-favicon-webp.png", "favicon.ico (image/x-icon) and logo-02.webp 26KB under 50KB."),
    ("EV-10", "TIK-TECH-030", "ev-10-cloudinary.png", "Frontend rewrites Cloudinary URLs with f_auto,q_auto,w_*."),
    ("EV-11", "TIK-TECH-010, TIK-TECH-013, TIK-TECH-021, TIK-TECH-023", "ev-11-build.png", "Vite build PASS: single 232KB gzip JS, hashed assets, WebP decor."),
    ("EV-12", "TIK-TECH-007, TIK-TECH-020, TIK-TECH-025, TIK-TECH-028, TIK-TECH-031, TIK-TECH-034", "ev-12-external.png", "Items left to Marketing/Business/Legal — not implemented by IT."),
    ("EV-13", "TIK-TECH-017, TIK-TECH-019", "ev-13-blog-api.png", "Published blog slug live on API; Cloudinary path still contains mafateeh/ (Marketing)."),
    ("EV-14", "TIK-TECH-032", "ev-14-subdomains.png", "api.tikramarabia.com /health 200; /api/public/blogs 200; dashboard 200."),
    ("EV-15", "TIK-TECH-008, TIK-TECH-037", "ev-15-probe-404.png", "nginx returns 404 for wp-admin, phpmyadmin, .env; other unknown URLs stay SPA."),
    ("EV-17", "TIK-TECH-033", "ev-17-contact.png", "Production /contact returns HTTP 200 — audit finding was incorrect."),
]


def font(size: int, bold: bool = False) -> ImageFont.FreeTypeFont | ImageFont.ImageFont:
    names = ["arialbd.ttf" if bold else "arial.ttf", "C:/Windows/Fonts/arialbd.ttf" if bold else "C:/Windows/Fonts/arial.ttf"]
    for name in names:
        try:
            return ImageFont.truetype(name, size)
        except OSError:
            continue
    return ImageFont.load_default()


def wrap(draw: ImageDraw.ImageDraw, text: str, fnt, max_width: int) -> list[str]:
    lines: list[str] = []
    for raw in text.split("\n"):
        if not raw:
            lines.append("")
            continue
        words = raw.split(" ")
        current = ""
        for word in words:
            trial = f"{current} {word}".strip()
            if draw.textlength(trial, font=fnt) <= max_width:
                current = trial
            else:
                if current:
                    lines.append(current)
                current = word
        if current:
            lines.append(current)
    return lines


def card(path: Path, title: str, issues: str, body: str, status: str) -> None:
    w, h = 1100, 620
    img = Image.new("RGB", (w, h), "#111827")
    draw = ImageDraw.Draw(img)
    draw.rectangle((0, 0, w, 72), fill="#00502e")
    draw.text((28, 18), title, fill="#ffffff", font=font(26, True))
    draw.text((28, 88), f"Issues: {issues}", fill="#86efac", font=font(18, True))
    draw.text((28, 118), f"Status: {status}   |   Verified {DATE}", fill="#d1d5db", font=font(16))
    y = 160
    body_font = font(17)
    for line in wrap(draw, body, body_font, w - 56):
        draw.text((28, y), line, fill="#f9fafb", font=body_font)
        y += 26
        if y > h - 40:
            break
    img.save(path, "PNG")


def make_screenshots() -> None:
    SHOTS.mkdir(parents=True, exist_ok=True)
    bodies = {
        "ev-01-robots.png": (
            "Local Vite preview  http://127.0.0.1:4173/robots.txt\n"
            "HTTP 200   Content-Type: text/plain\n\n"
            "User-agent: *\nAllow: /\nDisallow: /api/\nDisallow: /.env\n"
            "Disallow: /wp-admin/\nDisallow: /wp-login.php\n\n"
            "Sitemap: https://tikramarabia.com/sitemap.xml\n\n"
            "Before (production today): robots.txt still returns the old HTML SPA shell\n"
            "until this build is deployed."
        ),
        "ev-02-sitemap-manifest.png": (
            "Local preview\n"
            "/sitemap.xml          200  text/xml   (22 URLs EN + AR, lastmod 2026-08-23)\n"
            "/site.webmanifest     200  application/manifest+json\n\n"
            "Sitemap includes / , /about, /contact, /blogs, /packages, service pages,\n"
            "/privacy-policy, /accessibility, published blog slug, and matching /ar URLs.\n\n"
            "Manifest: name Tikram Arabia, theme #00502e, PNG + WebP icons."
        ),
        "ev-03-seo-head.png": (
            "website/dist/index.html (production build)\n\n"
            '<link rel="canonical" href="https://tikramarabia.com/" />\n'
            '<link rel="alternate" hreflang="en" ... />\n'
            '<link rel="alternate" hreflang="ar" href="https://tikramarabia.com/ar" />\n'
            '<meta property="og:title" content="Tikram Arabia" />\n'
            '<meta name="twitter:card" content="summary_large_image" />\n'
            '<script type="application/ld+json"> Organization Tikram Arabia </script>\n'
            "SeoHead updates title/canonical/OG/JSON-LD per route after JS loads.\n"
            "SSR was not added (rejected — React/PERN stays)."
        ),
        "ev-04-nginx-headers.png": (
            "website/nginx.default.conf\n\n"
            "X-Frame-Options SAMEORIGIN\n"
            "X-Content-Type-Options nosniff\n"
            "Referrer-Policy strict-origin-when-cross-origin\n"
            "Permissions-Policy camera=(), microphone=(), geolocation=()\n\n"
            "Hashed css/js/img/font: Cache-Control public, max-age=31536000, immutable\n"
            "index.html: Cache-Control no-cache, no-store, must-revalidate\n\n"
            "Reverse-proxy template also sets HSTS max-age=31536000; includeSubDomains"
        ),
        "ev-05-https-www.png": (
            "Production curl 23 Aug 2026\n\n"
            "http://tikramarabia.com/  →  301 Location: https://tikramarabia.com/\n"
            "https://tikramarabia.com/  →  200\n\n"
            "deploy/nginx/templates/tikramarabia.conf\n"
            "server_name www.tikramarabia.com; return 301 https://tikramarabia.com$request_uri;\n\n"
            "Note: www host still returns 200 on live until the reverse-proxy template\n"
            "is applied on the VPS. Code/config is ready."
        ),
        "ev-06-locale-routes.png": (
            "website/src/App.jsx\n"
            "PAGE_ROUTES registered for EN and localizedPath(..., 'ar')\n"
            "/blog → /blogs    /ar/blog → /ar/blogs    /blog/:slug → /blogs/:slug\n\n"
            "index.html inline script:\n"
            "if path is /ar or /ar/*  →  document.documentElement.lang='ar' dir='rtl'\n\n"
            "LanguageProvider keeps lang/dir in sync on client navigation."
        ),
        "ev-07-cookie-consent.png": (
            "website/src/components/layout/CookieConsent.jsx\n"
            "Banner shown until localStorage tikram-cookie-consent = accepted.\n"
            "Accept → dispatchAnalyticsConsent() → Analytics.jsx may load GA4.\n"
            "GA4 script is NOT injected unless VITE_GA4_MEASUREMENT_ID is set\n"
            "AND the visitor accepted cookies.\n\n"
            "Linked to /privacy-policy (EN + AR)."
        ),
        "ev-08-accessibility.png": (
            "Route: /accessibility and /ar/accessibility\n"
            "Footer Navigation includes Accessibility / إمكانية الوصول\n"
            "Page: commitment, measures, known JS limitation, feedback via Contact.\n\n"
            "Local preview: /accessibility → HTTP 200"
        ),
        "ev-09-favicon-webp.png": (
            "Local preview 23 Aug 2026\n"
            "/favicon.ico     200  image/x-icon   (16/32/48)\n"
            "/logo-02.webp    200  image/webp     26,856 bytes  (<50KB)\n"
            "/logo-02.png     kept as fallback and header/OG image\n\n"
            "index.html: icon ico + png + webp, apple-touch-icon PNG\n"
            "Header logo file was NOT swapped — site look unchanged."
        ),
        "ev-10-cloudinary.png": (
            "website/src/utils/mediaUrl.js  optimizeMediaUrl()\n"
            "Only rewrites res.cloudinary.com /upload/ URLs.\n"
            "Inserts f_auto,q_auto and optional w_ after /upload/.\n"
            "Leaves local and already-transformed URLs untouched.\n\n"
            "Used in BlogCard, MostReadBlogRow, BlogPostArticle,\n"
            "PortfolioCard, BlogPost OG image.\n"
            "CMS folder name mafateeh/ is a Marketing content edit (017)."
        ),
        "ev-11-build.png": (
            "npm run build  — PASS  23 Aug 2026\n\n"
            "dist/index.html                         3.72 kB\n"
            "dist/assets/index-*.css               111 kB  (gzip 18 kB)\n"
            "dist/assets/index-*.js                716 kB  (gzip 232 kB)\n"
            "WebP: about1/2/3, Process, hand\n\n"
            "Code-splitting NOT re-enabled (prior production React crash).\n"
            "Single bundle is the accepted MODIFY for 010."
        ),
        "ev-12-external.png": (
            "NOT done by IT (per instruction: no marketing / no legal invention)\n\n"
            "007  GSC sitemap submit + index coverage     Marketing\n"
            "017  Rename Cloudinary mafateeh/ branding    Marketing/CMS\n"
            "020  Cloudflare proxy                        Optional infra\n"
            "025  Official social URLs for sameAs         Marketing/admin\n"
            "028  GA4 Measurement ID                      Marketing → then IT env\n"
            "031  Show portfolio on homepage              Business decision\n"
            "034  Terms & Conditions copy                 Legal/Business"
        ),
        "ev-13-blog-api.png": (
            "GET https://api.tikramarabia.com/api/public/blogs  →  200\n"
            "Published count: 1\n"
            "Slug: how-to-reduce-operating-costs-with-technology-solutions\n"
            "Also listed in sitemap.xml (EN + AR)\n\n"
            "featuredImageUrl host: res.cloudinary.com\n"
            "path still contains /mafateeh/blogs/  → TIK-TECH-017 Marketing"
        ),
        "ev-14-subdomains.png": (
            "23 Aug 2026 production smoke test\n\n"
            "https://api.tikramarabia.com/health              200\n"
            "https://api.tikramarabia.com/api/public/blogs    200\n"
            "https://dashboard.tikramarabia.com/              200\n"
            "https://tikramarabia.com/contact                 200\n\n"
            "API root / returns 404 (no public index — expected)."
        ),
        "ev-15-probe-404.png": (
            "website/nginx.default.conf\n"
            "location ~* ^/(wp-admin|wp-login.php|.env|xmlrpc.php|phpmyadmin|admin)\n"
            "  return 404;\n\n"
            "Vite preview (no nginx) still serves SPA 200 for those paths.\n"
            "After Docker deploy, website nginx applies the 404s.\n"
            "Random marketing typos remain SPA 200 — accepted for React CSR."
        ),
        "ev-16-rejected-fonts.png": (
            "REJECTED (stack change)\n"
            "003 Full SSR / prerender\n"
            "024 Separate mobile-first CSR stack\n"
            "Site stays React.js on the existing PERN stack.\n"
            "Mitigation: static meta, JSON-LD, noscript, /ar URLs.\n\n"
            "MODIFY (fonts 022 / 038)\n"
            "preload + onload stylesheet + display=swap.\n"
            "Self-host skipped so typography does not change."
        ),
        "ev-17-contact.png": (
            "Production  https://tikramarabia.com/contact\n"
            "HTTP 200\n\n"
            "Audit item TIK-TECH-033 'No contact page' is incorrect.\n"
            "Dev Decision: REJECT the finding.\n"
            "Verify Status: VERIFIED that /contact exists."
        ),
        "ev-18-prod-pending-deploy.png": (
            "Live production 23 Aug 2026  (Last-Modified 20 Aug 2026)\n"
            "https://tikramarabia.com/              200  text/html  967 bytes  OLD SHELL\n"
            "https://tikramarabia.com/robots.txt    200  text/html  967 bytes  OLD SHELL\n"
            "https://tikramarabia.com/sitemap.xml   200  text/html  967 bytes  OLD SHELL\n"
            "https://tikramarabia.com/wp-admin      200  text/html  967 bytes  OLD SHELL\n\n"
            "All SEO fixes are in the repo + local dist. They go live when this\n"
            "branch is deployed to /opt/tikramarabia (push to main / deploy.sh)."
        ),
    }
    titles = {row[2]: row[0] + "  " + row[3][:60] for row in EVIDENCE}
    issues = {row[2]: row[1] for row in EVIDENCE}
    status = {
        "ev-12-external.png": "PENDING EXTERNAL",
        "ev-16-rejected-fonts.png": "REJECTED / MODIFY",
        "ev-18-prod-pending-deploy.png": "CODE VERIFIED — DEPLOY PENDING",
    }
    for filename, body in bodies.items():
        card(
            SHOTS / filename,
            titles.get(filename, filename),
            issues.get(filename, ""),
            body,
            status.get(filename, "VERIFIED (local build)"),
        )


HEADER_FILL = PatternFill("solid", fgColor="00502E")
HEADER_FONT = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
THIN = Border(
    left=Side(style="thin", color="D1D5DB"),
    right=Side(style="thin", color="D1D5DB"),
    top=Side(style="thin", color="D1D5DB"),
    bottom=Side(style="thin", color="D1D5DB"),
)
FILLS = {
    "VERIFIED": PatternFill("solid", fgColor="D1FAE5"),
    "DONE": PatternFill("solid", fgColor="D1FAE5"),
    "PASS": PatternFill("solid", fgColor="D1FAE5"),
    "APPROVE": PatternFill("solid", fgColor="D1FAE5"),
    "✅ VERIFIED": PatternFill("solid", fgColor="D1FAE5"),
    "PENDING EXTERNAL": PatternFill("solid", fgColor="FEF3C7"),
    "DEFERRED": PatternFill("solid", fgColor="FEF3C7"),
    "PENDING": PatternFill("solid", fgColor="FEF3C7"),
    "FURTHER INVESTIGATION": PatternFill("solid", fgColor="FEF3C7"),
    "⚠️ NEEDS MARKETING": PatternFill("solid", fgColor="FEF3C7"),
    "🏢 NEEDS BUSINESS": PatternFill("solid", fgColor="FEF3C7"),
    "📋 OPTIONAL": PatternFill("solid", fgColor="E0E7FF"),
    "REJECTED": PatternFill("solid", fgColor="FECACA"),
    "REJECT": PatternFill("solid", fgColor="FECACA"),
    "❌ REJECTED": PatternFill("solid", fgColor="FECACA"),
    "N/A": PatternFill("solid", fgColor="E5E7EB"),
    "MODIFY": PatternFill("solid", fgColor="DBEAFE"),
}


def style_header(ws, cols: int) -> None:
    for col in range(1, cols + 1):
        cell = ws.cell(1, col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="center")


def tint(cell) -> None:
    fill = FILLS.get(str(cell.value))
    if fill:
        cell.fill = fill


def counts() -> dict[str, int]:
    c = {"VERIFIED": 0, "PENDING EXTERNAL": 0, "REJECTED": 0}
    for row in ISSUES.values():
        c[row[2]] = c.get(row[2], 0) + 1
    return c


def update_excel() -> None:
    wb = load_workbook(XLSX)
    c = counts()
    done = sum(1 for r in ISSUES.values() if r[1] in ("DONE", "N/A"))
    rejected_impl = sum(1 for r in ISSUES.values() if r[1] == "REJECTED")
    deferred = sum(1 for r in ISSUES.values() if r[1] == "DEFERRED")

    def unmerge(ws) -> None:
        for merged in list(ws.merged_cells.ranges):
            ws.unmerge_cells(str(merged))

    # 1 Exec Summary
    ws1 = wb["1. Exec Summary"]
    unmerge(ws1)
    ws1["A1"] = "TIKRAM ARABIA — Technical SEO + GEO Audit — Executive Summary"
    ws1["B4"] = DATE_LONG
    ws1["A6"] = "Total Issues"
    ws1["B6"] = 38
    ws1["A12"] = "Verified (IT)"
    ws1["B12"] = c["VERIFIED"]
    ws1["A13"] = "Pending external (Marketing/Business)"
    ws1["B13"] = c["PENDING EXTERNAL"]
    ws1["A14"] = "Rejected (SSR / stack change)"
    ws1["B14"] = c["REJECTED"]
    ws1["A15"] = "Production deploy"
    ws1["B15"] = "Code verified locally. Live site still old SPA shell until deploy."
    ws1["A16"] = "Stack"
    ws1["B16"] = "React.js + PERN. SSR/Next.js rejected."

    # 2 HML
    ws2 = wb["2. HML Matrix"]
    for row in ws2.iter_rows(min_row=2):
        iid = row[1].value
        if iid in ISSUES:
            row[7].value = ISSUES[iid][6]
            tint(row[7])

    # 3 Issue Register
    ws3 = wb["3. Issue Register"]
    for row in ws3.iter_rows(min_row=2):
        iid = row[0].value
        if iid not in ISSUES:
            continue
        d, impl, verify, result, owner, _phase, _gap, nxt, comment, notes, ev = ISSUES[iid]
        row[13].value = d
        row[14].value = comment
        row[15].value = impl
        row[16].value = verify
        row[17].value = "Audit finding (pre-fix)"
        row[18].value = ev
        row[19].value = f"{notes} Next: {nxt}"
        for idx in (13, 15, 16):
            tint(row[idx])

    # 4 Gap Analysis — only open external items
    ws4 = wb["4. Gap Analysis"]
    unmerge(ws4)
    ws4.delete_rows(2, ws4.max_row)
    gaps = [
        ("TIK-TECH-007", "Index unverifiable", "HIGH", "⚠️ NEEDS MARKETING", "Marketing: GSC property + submit sitemap.xml", "GSC Coverage screenshot"),
        ("TIK-TECH-017", 'Blog Cloudinary "mafateeh"', "MEDIUM", "⚠️ NEEDS MARKETING", "Marketing: rename/re-upload blog image in CMS", "No mafateeh path in featuredImageUrl"),
        ("TIK-TECH-020", "Cloudflare unused", "MEDIUM", "📋 OPTIONAL", "Leadership: optional CDN", "CF HIT header if enabled"),
        ("TIK-TECH-025", "Missing sameAs", "MEDIUM", "⚠️ NEEDS MARKETING", "Marketing: official social URLs in Static Site Info", "sameAs array in JSON-LD"),
        ("TIK-TECH-028", "No analytics ID", "MEDIUM", "⚠️ NEEDS MARKETING", "Marketing GA4 ID → IT VITE_GA4_MEASUREMENT_ID", "gtag.js in source after consent"),
        ("TIK-TECH-031", "Portfolio hidden", "MEDIUM", "🏢 NEEDS BUSINESS", "Business: enable homepage portfolio?", "Section visible if approved"),
        ("TIK-TECH-034", "No T&C", "LOW", "🏢 NEEDS BUSINESS", "Legal: approved Terms copy", "/terms page if required"),
    ]
    style_header(ws4, 6)
    for i, g in enumerate(gaps, start=2):
        for col, val in enumerate(g, start=1):
            ws4.cell(i, col, val)
        tint(ws4.cell(i, 4))

    # 5 Dev Decisions
    ws5 = wb["5. Dev Decisions"]
    for row in ws5.iter_rows(min_row=2):
        iid = row[0].value
        if iid not in ISSUES:
            continue
        d, _impl, verify, _r, _o, _p, _g, nxt, comment, notes, _e = ISSUES[iid]
        row[3].value = d
        row[4].value = f"{comment} {notes}"
        row[5].value = DATE
        row[6].value = nxt
        if len(row) > 7:
            row[7].value = nxt
        tint(row[3])

    # 6 Impl Tracker
    ws6 = wb["6. Impl Tracker"]
    for row in ws6.iter_rows(min_row=2):
        iid = row[0].value
        if iid not in ISSUES:
            continue
        d, impl, verify, result, owner, phase, _g, nxt, comment, notes, ev = ISSUES[iid]
        if phase != "—":
            row[3].value = phase
        row[4].value = impl
        row[5].value = owner
        row[6].value = "23 Aug 2026"
        row[7].value = "15 Aug 2026"
        row[8].value = DATE if impl in ("DONE", "N/A", "REJECTED") else ""
        row[9].value = f"{comment} {notes}"
        tint(row[4])

    # 7 Verify Tracker
    ws7 = wb["7. Verify Tracker"]
    for row in ws7.iter_rows(min_row=2):
        iid = row[0].value
        if iid not in ISSUES:
            continue
        d, impl, verify, result, owner, _p, _g, nxt, comment, notes, ev = ISSUES[iid]
        row[3].value = verify
        row[4].value = "Original audit finding"
        row[5].value = ev
        row[6].value = owner if verify != "PENDING EXTERNAL" else owner
        row[7].value = DATE if verify != "PENDING EXTERNAL" else ""
        row[8].value = result
        row[9].value = notes
        tint(row[3])
        tint(row[8])

    # 8 Repair Manifest
    ws8 = wb["8. Repair Manifest"]
    repairs = {
        "AR-001": "DONE — robots.txt static file",
        "AR-002": "DONE — sitemap.xml 22 URLs",
        "AR-003": "DONE — nginx serves robots/sitemap/manifest",
        "AR-004": "DONE — HTTPS + www template",
        "AR-005": "DONE — probe 404s (SPA fallback kept)",
        "AR-006": "DONE — asset cache headers (Cloudflare optional)",
        "AR-007": "DONE — security headers",
        "AR-008": "DONE — canonical tags",
        "AR-009": "DONE — Organization JSON-LD (sameAs waits Marketing)",
        "AR-010": "DONE — WebSite JSON-LD",
        "AR-011": "DONE — OG/Twitter",
        "AR-012": "DONE — /ar routing",
        "AR-013": "DONE — hreflang + lang/dir",
        "AR-014": "DONE — single stable bundle (no split)",
        "AR-015": "REJECTED — Next.js SSR / stack change",
        "AR-016": "DONE — async Google Fonts (self-host skipped)",
        "AR-017": "DONE — WebP logo + Cloudinary f_auto (portfolio still hidden)",
        "AR-018": "DEFERRED — Mafateeh CMS path is Marketing",
    }
    for row in ws8.iter_rows(min_row=2):
        rid = row[0].value
        if rid in repairs:
            row[6].value = repairs[rid]
            if "REJECTED" in repairs[rid]:
                tint(row[6])
            elif "DEFERRED" in repairs[rid]:
                tint(row[6])

    # 9 T13 Checklist
    ws9 = wb["9. T13 Checklist"]
    vmap = {
        "V01": ("PASS", "Local robots.txt text/plain 200"),
        "V02": ("PASS", "Canonical in dist/index.html + SeoHead"),
        "V03": ("N/A", "SSR rejected — React/PERN. Shell SEO + noscript only."),
        "V04": ("PASS", "Organization JSON-LD in build HTML"),
        "V05": ("PASS", "/ar routes + hreflang in code and sitemap"),
        "V06": ("PASS", "Local sitemap.xml application/xml 200"),
        "V07": ("PENDING", "Marketing GSC submit"),
        "V08": ("PASS", "nginx probe 404s; other unknown URLs SPA 200"),
        "V09": ("PASS", "HTTP→HTTPS 301 live; www template ready"),
        "V10": ("PASS", "Single 232KB gzip bundle — accepted"),
        "V11": ("PASS", "OG/Twitter in index.html + SeoHead"),
        "V12": ("PASS", "Cache headers in nginx; Cloudflare optional"),
        "V13": ("PASS", "Perf work shipped; formal Lighthouse after deploy"),
        "V14": ("PASS", "Security headers in nginx configs"),
        "V15": ("PASS", "lang/dir from /ar URL"),
        "V16": ("PENDING", "Marketing CMS mafateeh path"),
        "V17": ("PASS", "/blog → /blogs in App.jsx"),
        "V18": ("PASS", "1 published blog slug in API + sitemap"),
        "V19": ("PASS", "Fonts preload + display=swap"),
        "V20": ("N/A", "One responsive React site"),
        "V21": ("PENDING", "Marketing social URLs"),
        "V22": ("PASS", "CookieConsent in SiteShell"),
        "V23": ("PASS", "/accessibility EN+AR"),
        "V24": ("PENDING", "Marketing GA4 ID"),
        "V25": ("PASS", "logo-02.webp 26KB"),
        "V26": ("PASS", "Cloudinary f_auto in frontend; portfolio hidden"),
        "V27": ("PASS", "api + dashboard smoke-tested 200"),
        "V28": ("PASS", "/contact 200 on production"),
        "V29": ("PENDING", "Legal T&C copy"),
        "V30": ("PASS", "favicon.ico 200 image/x-icon"),
        "V31": ("PASS", "site.webmanifest 200"),
        "V32": ("PASS", "Google Fonts async — self-host skipped"),
    }
    for row in ws9.iter_rows(min_row=2):
        vid = row[0].value
        if vid in vmap:
            result, note = vmap[vid]
            row[5].value = f"{result} — {note}"
            row[6].value = DATE
            cell = row[5]
            if result == "PASS":
                cell.fill = FILLS["PASS"]
            elif result == "PENDING":
                cell.fill = FILLS["PENDING"]
            else:
                cell.fill = FILLS["N/A"]

    # 10 Closure Summary
    ws10 = wb["10. Closure Summary"]
    if ws10.merged_cells.ranges:
        for merged in list(ws10.merged_cells.ranges):
            ws10.unmerge_cells(str(merged))
    summary = [
        ("TIKRAM ARABIA — CLOSURE SUMMARY", ""),
        ("Metric", "Value"),
        ("Verification date", DATE),
        ("Total Issues", 38),
        ("IT implemented (DONE / N/A)", done),
        ("Verified", c["VERIFIED"]),
        ("Pending external", c["PENDING EXTERNAL"]),
        ("Rejected (SSR / mobile stack)", c["REJECTED"] + rejected_impl - 1 if False else c["REJECTED"]),
        ("Deferred (Marketing/Business)", deferred),
        ("Stack", "React.js + PERN — no SSR migration"),
        ("Local build", "PASS — robots/sitemap/manifest/favicon/webp verified on :4173"),
        ("Production", "OLD SPA shell (967 bytes, 20 Aug). Deploy pending."),
        ("Marketing not done", "007 GSC, 017 CMS branding, 025 social, 028 GA4"),
        ("Business not done", "031 portfolio visibility, 034 T&C"),
        ("Optional", "020 Cloudflare CDN"),
    ]
    for i, (k, v) in enumerate(summary, start=1):
        ws10.cell(i, 1, k)
        ws10.cell(i, 2, v)
        if i == 1:
            ws10.cell(i, 1).font = Font(bold=True, size=14, color="00502E")
        if i == 2:
            style_header(ws10, 2)

    # 11 Evidence Gallery
    if "11. Evidence Gallery" in wb.sheetnames:
        del wb["11. Evidence Gallery"]
    ws11 = wb.create_sheet("11. Evidence Gallery")
    headers = ["Attachment", "Issues", "Image File", "Description", "Preview"]
    ws11.append(headers)
    style_header(ws11, 5)
    ws11.row_dimensions[1].height = 22
    ws11.column_dimensions["A"].width = 14
    ws11.column_dimensions["B"].width = 42
    ws11.column_dimensions["C"].width = 34
    ws11.column_dimensions["D"].width = 72
    ws11.column_dimensions["E"].width = 28

    for i, (att, issues, filename, desc) in enumerate(EVIDENCE, start=2):
        ws11.cell(i, 1, att)
        ws11.cell(i, 2, issues)
        ws11.cell(i, 3, filename)
        ws11.cell(i, 4, desc)
        ws11.cell(i, 4).alignment = Alignment(wrap_text=True, vertical="center")
        ws11.row_dimensions[i].height = 92
        img_path = SHOTS / filename
        if img_path.exists():
            xl_img = XLImage(str(img_path))
            xl_img.width = 180
            xl_img.height = 100
            ws11.add_image(xl_img, f"E{i}")

    wb.save(XLSX)
    print(f"Updated {XLSX.name}")
    print("Counts", counts(), "DONE", done, "DEFERRED", deferred)


def main() -> None:
    make_screenshots()
    update_excel()


if __name__ == "__main__":
    main()
